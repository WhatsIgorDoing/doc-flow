"""
Serviço de preenchimento de templates Excel.
Implementa ITemplateFiller de forma assíncrona.
"""

import asyncio
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.core.interfaces import (
    IFileSystemManager,
    ITemplateFiller,
)
from app.core.logger import app_logger
from app.domain.entities import DocumentGroup
from app.domain.exceptions import (
    TemplateFillError,
    TemplateNotFoundError,
)
from app.domain.file_naming import get_filename_with_revision


class OpenpyxlTemplateFiller(ITemplateFiller):
    """Implementação que usa openpyxl para preencher templates Excel."""

    def __init__(self, file_manager: IFileSystemManager):
        self._file_manager = file_manager

    async def fill_and_save(
        self, template_path: Path, output_path: Path, data: List[DocumentGroup]
    ) -> None:
        """
        Preenche um template Excel com os dados dos documentos.
        Executa operações pesadas de Excel em thread pool.
        """
        if not template_path.exists():
            app_logger.error("Template not found", extra={"path": str(template_path)})
            # Levanta a exceção diretamente para ser tratada acima
            raise TemplateNotFoundError(
                f"Arquivo template não encontrado: {template_path}"
            )

        try:
            # 1. Copia o template para o local de saída (async)
            await self._file_manager.copy_file(template_path, output_path)

            # 2. Abre e preenche em executor (CPU/IO intensive)
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(
                None, self._fill_workbook_sync, output_path, data
            )

            app_logger.info(
                "Template filled successfully",
                extra={"output_path": str(output_path), "groups_count": len(data)},
            )

        except (TemplateNotFoundError, TemplateFillError):
            raise
        except Exception as e:
            app_logger.error(
                "Failed to fill template",
                extra={"output_path": str(output_path), "error": str(e)},
            )
            raise TemplateFillError(f"Falha ao preencher o template {output_path}: {e}")

    def _fill_workbook_sync(self, output_path: Path, data: List[DocumentGroup]) -> None:
        """Lógica síncrona de preenchimento do Excel."""
        workbook = openpyxl.load_workbook(output_path)
        sheet: Worksheet = workbook.active

        # Encontrar onde inserir os dados (após cabeçalho, antes de "FIM")
        insert_row = 2
        for row_num in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=1).value
            if cell_value and str(cell_value).upper() == "FIM":
                insert_row = row_num
                break

        # Preparar dados
        all_rows_data = []
        for group in data:
            manifest_info = group.files[0].associated_manifest_item
            if not manifest_info:
                continue

            for file in group.files:
                revision = manifest_info.revision
                filename_with_revision = get_filename_with_revision(
                    file.path.name, revision
                )

                row_data = [
                    manifest_info.document_code,
                    manifest_info.revision,
                    manifest_info.title,
                    filename_with_revision,
                    manifest_info.metadata.get("FORMATO", "A4"),
                    manifest_info.metadata.get("DISCIPLINA", ""),
                    manifest_info.metadata.get("TIPO DE DOCUMENTO", ""),
                    manifest_info.metadata.get("PROPÓSITO", ""),
                    manifest_info.metadata.get("CAMINHO DATABOOK", ""),
                ]
                all_rows_data.append(row_data)

        # Inserir e Preencher
        if all_rows_data:
            # Insere linhas (em lote para performance)
            sheet.insert_rows(insert_row, amount=len(all_rows_data))

            # Preenche dados
            for i, row_data in enumerate(all_rows_data):
                target_row = insert_row + i
                for col_num, value in enumerate(row_data, 1):
                    sheet.cell(row=target_row, column=col_num, value=value)

        # Formatação
        self._apply_formatting(sheet, insert_row, len(all_rows_data))

        workbook.save(output_path)
        workbook.close()

    def _apply_formatting(
        self, sheet: Worksheet, data_start_row: int, num_data_rows: int
    ):
        """Aplica formatação básica."""
        num_columns = 9

        # Helper para borda fina
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Cabeçalho (amarelo)
        header_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        for col in range(1, num_columns + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dados
        for r in range(data_start_row, data_start_row + num_data_rows):
            for c in range(1, num_columns + 1):
                cell = sheet.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Ajustar larguras
        widths = {
            "A": 35,
            "B": 10,
            "C": 60,
            "D": 35,
            "E": 10,
            "F": 20,
            "G": 20,
            "H": 20,
            "I": 20,
        }
        for col_letter, width in widths.items():
            sheet.column_dimensions[col_letter].width = width
