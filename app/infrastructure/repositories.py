"""
Repositórios assíncronos para acesso a dados.
Implementa padrão Repository com I/O não bloqueante.
"""

import asyncio
import os
from pathlib import Path
from typing import List

import aiofiles
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.core.interfaces import IFileRepository, IFileSystemManager, IManifestRepository
from app.core.logger import app_logger
from app.domain.entities import DocumentFile, ManifestItem
from app.domain.exceptions import (
    FileOperationError,
    FileReadError,
    ManifestParseError,
    ManifestReadError,
    SourceDirectoryNotFoundError,
)


class ManifestRepository(IManifestRepository):
    """Repositório para leitura de manifestos Excel de forma assíncrona."""

    async def load_from_file(self, file_path: Path) -> List[ManifestItem]:
        """
        Carrega itens do manifesto de um arquivo Excel.

        Args:
            file_path: Caminho para o arquivo Excel do manifesto

        Returns:
            Lista de itens do manifesto

        Raises:
            ManifestReadError: Se o arquivo não existir ou houver erro de leitura
            ManifestParseError: Se houver erro ao parsear os dados
        """
        if not file_path.exists():
            app_logger.error(
                "Manifest file not found", extra={"file_path": str(file_path)}
            )
            raise ManifestReadError(f"Manifest file not found: {file_path}")

        try:
            # openpyxl não é nativo async, então rodamos em executor
            loop = asyncio.get_event_loop()
            items = await loop.run_in_executor(None, self._read_excel_sync, file_path)

            app_logger.info(
                "Manifest loaded successfully",
                extra={"file_path": str(file_path), "items_count": len(items)},
            )
            return items

        except ManifestReadError:
            raise
        except Exception as e:
            app_logger.error(
                "Error parsing manifest",
                extra={"file_path": str(file_path), "error": str(e)},
            )
            raise ManifestParseError(f"Error parsing manifest: {e}")

    def _read_excel_sync(self, file_path: Path) -> List[ManifestItem]:
        """
        Leitura síncrona do Excel (executada em thread pool).
        Suporta formato dinâmico detectando cabeçalho.

        Args:
            file_path: Caminho do arquivo Excel

        Returns:
            Lista de itens do manifesto
        """
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet: Worksheet = workbook.active

        items: List[ManifestItem] = []

        # Configurações de detecção
        MAX_HEADER_SCAN = 20
        header_row_idx = 0

        # Mapa de colunas (nome -> índice)
        col_map = {}

        # Palavras-chave para identificar colunas (lower case)
        KEYWORDS = {
            "document_code": ["documento", "código", "codigo", "code", "doc"],
            "revision": ["revisão", "revisao", "rev", "revision"],
            "title": ["título", "titulo", "title", "descrição", "descricao"],
        }

        # 1. Detectar linha de cabeçalho (Melhor Match)
        best_header_row = 0
        best_col_map = {}
        max_matches = 0

        for i, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN, values_only=True), 1
        ):
            row_values = [str(cell).lower().strip() if cell else "" for cell in row]

            # Tenta encontrar colunas obrigatórias
            found_cols = {}
            for target_field, keywords in KEYWORDS.items():
                for idx, cell_res in enumerate(row_values):
                    if any(k == cell_res or k in cell_res for k in keywords):
                        # Evitar sobrescrever se já encontrou match exato/melhor
                        if target_field not in found_cols:
                            found_cols[target_field] = idx
                        elif cell_res in keywords:  # Prioridade para match exato
                            found_cols[target_field] = idx

            # Pontuação: Quantas colunas obrigatórias encontrou
            match_count = len(found_cols)

            # Critério de seleção:
            # 1. Deve ter 'document_code'
            # 2. Deve ter mais matches que o anterior
            # 3. Se empate, prefere o que tem mais matches exatos (opcional, aqui simplificado)
            if "document_code" in found_cols and match_count >= 2:
                if match_count > max_matches:
                    max_matches = match_count
                    best_header_row = i
                    best_col_map = found_cols

        if best_header_row > 0:
            header_row_idx = best_header_row
            col_map = best_col_map
            app_logger.info(
                f"Header detected at row {header_row_idx}. Map: {col_map} (Matches: {max_matches})"
            )

        # Fallback: Se não detectou cabeçalho, assume formato legado (A=Code, B=Rev, C=Title, Row 1=Header)
        if not header_row_idx:
            header_row_idx = 1
            col_map = {"document_code": 0, "revision": 1, "title": 2}
            app_logger.warning(
                "Header not detected dynamically, using legacy fallback (A=Code, B=Rev, C=Title)"
            )

        # Obter lista de cabeçalhos para metadados
        # Precisamos ler a linha de cabeçalho novamente para ter os nomes originais
        header_cells = []
        for row in sheet.iter_rows(
            min_row=header_row_idx, max_row=header_row_idx, values_only=True
        ):
            header_cells = [
                str(cell).strip() if cell else f"Column_{i}"
                for i, cell in enumerate(row)
            ]
            break

        # 2. Ler dados
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            # Ignora linhas totalmente vazias
            if not any(row):
                continue

            # Extrair campos principais usando o mapa
            try:
                # Code
                idx_code = col_map.get("document_code")
                raw_code = (
                    row[idx_code]
                    if idx_code is not None and idx_code < len(row)
                    else None
                )
                document_code = str(raw_code).strip() if raw_code else ""

                # Ignora se não tiver código (linha inválida ou de formatação)
                if not document_code or len(document_code) < 3:
                    continue

                # Validar se é um código real (ex: ignorar linhas que repetem cabeçalho ou são notas)
                # Heurística simples: não deve ser igual ao nome da coluna
                # E também não deve ser igual keywords de outras colunas (ex: "Revisão")
                is_header_repeat = False
                doc_lower = document_code.lower()
                for k_list in KEYWORDS.values():
                    if doc_lower in k_list:
                        is_header_repeat = True
                        break

                if is_header_repeat:
                    continue

                # Revision
                idx_rev = col_map.get("revision")
                raw_rev = (
                    row[idx_rev] if idx_rev is not None and idx_rev < len(row) else None
                )
                revision = str(raw_rev).strip() if raw_rev else "0"

                # Title
                idx_title = col_map.get("title")
                raw_title = (
                    row[idx_title]
                    if idx_title is not None and idx_title < len(row)
                    else None
                )
                title = str(raw_title).strip() if raw_title else ""

                # Metadata: tudo que não é coluna principal mapeada
                metadata = {}
                main_indices = set(col_map.values())
                for i, val in enumerate(row):
                    if (
                        i not in main_indices
                        and i < len(header_cells)
                        and val is not None
                    ):
                        metadata[header_cells[i]] = val

                items.append(
                    ManifestItem(
                        document_code=document_code,
                        revision=revision,
                        title=title,
                        metadata=metadata,
                    )
                )

            except Exception as e:
                # Logar e continuar (best effort)
                app_logger.warning(
                    f"Error parsing row in manifest: {e}",
                    extra={"row_content": str(row)[:100]},
                )
                continue

        workbook.close()
        return items


class FileRepository(IFileRepository):
    """Repositório para operações com arquivos do sistema de forma assíncrona."""

    async def list_files(self, directory: Path) -> List[DocumentFile]:
        """
        Lista todos os arquivos em um diretório recursivamente.

        Args:
            directory: Caminho do diretório a ser escaneado

        Returns:
            Lista de arquivos encontrados

        Raises:
            SourceDirectoryNotFoundError: Se o diretório não existir
            FileReadError: Se houver erro ao ler os arquivos
        """
        if not directory.exists():
            app_logger.error(
                "Source directory not found", extra={"directory": str(directory)}
            )
            raise SourceDirectoryNotFoundError(f"Directory not found: {directory}")

        if not directory.is_dir():
            app_logger.error("Path is not a directory", extra={"path": str(directory)})
            raise SourceDirectoryNotFoundError(f"Not a directory: {directory}")

        try:
            # Listagem de arquivos em thread pool (operação I/O)
            loop = asyncio.get_event_loop()
            files = await loop.run_in_executor(None, self._list_files_sync, directory)

            app_logger.info(
                "Files listed successfully",
                extra={"directory": str(directory), "files_count": len(files)},
            )
            return files

        except (SourceDirectoryNotFoundError, FileReadError):
            raise
        except Exception as e:
            app_logger.error(
                "Error listing files",
                extra={"directory": str(directory), "error": str(e)},
            )
            raise FileReadError(f"Error listing files: {e}")

    def _list_files_sync(self, directory: Path) -> List[DocumentFile]:
        """
        Listagem síncrona de arquivos (executada em thread pool).

        Args:
            directory: Diretório a ser escaneado

        Returns:
            Lista de DocumentFile
        """
        found_files: List[DocumentFile] = []

        # .rglob('*') busca recursivamente por todos os arquivos
        for path in directory.rglob("*"):
            if path.is_file():
                try:
                    size_bytes = path.stat().st_size
                    found_files.append(DocumentFile(path=path, size_bytes=size_bytes))
                except OSError as e:
                    app_logger.warning(
                        "Could not read file stats",
                        extra={"file": str(path), "error": str(e)},
                    )
                    continue

        return found_files


class FileSystemManager(IFileSystemManager):
    """Gerenciador de operações do sistema de arquivos de forma assíncrona."""

    async def create_directory(self, path: Path) -> None:
        """
        Cria um diretório e todos os pais necessários.

        Args:
            path: Caminho do diretório a ser criado

        Raises:
            FileOperationError: Se houver erro ao criar o diretório
        """
        try:
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(
                None, path.mkdir, True, True
            )  # parents=True, exist_ok=True

            app_logger.debug("Directory created", extra={"directory": str(path)})

        except Exception as e:
            app_logger.error(
                "Failed to create directory",
                extra={"directory": str(path), "error": str(e)},
            )
            raise FileOperationError(f"Failed to create directory {path}: {e}")

    async def move_file(self, source: Path, destination: Path) -> None:
        """
        Move um arquivo de origem para destino.

        Args:
            source: Caminho do arquivo de origem
            destination: Caminho de destino

        Raises:
            FileOperationError: Se houver erro ao mover o arquivo
        """
        try:
            # Garante que o diretório de destino exista
            await self.create_directory(destination.parent)

            # Move o arquivo em thread pool
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(None, source.rename, destination)

            app_logger.debug(
                "File moved",
                extra={"source": str(source), "destination": str(destination)},
            )

        except Exception as e:
            app_logger.error(
                "Failed to move file",
                extra={
                    "source": str(source),
                    "destination": str(destination),
                    "error": str(e),
                },
            )
            raise FileOperationError(f"Failed to move {source} to {destination}: {e}")

    async def copy_file(self, source: Path, destination: Path) -> None:
        """
        Copia um arquivo usando aiofiles para I/O assíncrono.

        Args:
            source: Caminho do arquivo de origem
            destination: Caminho de destino

        Raises:
            FileOperationError: Se houver erro ao copiar o arquivo
        """
        try:
            # Garante que o diretório de destino exista
            await self.create_directory(destination.parent)

            # Copia arquivo de forma assíncrona
            async with aiofiles.open(source, "rb") as src:
                async with aiofiles.open(destination, "wb") as dst:
                    while chunk := await src.read(1024 * 1024):  # 1MB chunks
                        await dst.write(chunk)

            app_logger.debug(
                "File copied",
                extra={"source": str(source), "destination": str(destination)},
            )

        except Exception as e:
            app_logger.error(
                "Failed to copy file",
                extra={
                    "source": str(source),
                    "destination": str(destination),
                    "error": str(e),
                },
            )
            raise FileOperationError(f"Failed to copy {source} to {destination}: {e}")

    async def rename_file(self, source: Path, new_name: str) -> Path:
        """
        Renomeia um arquivo de forma segura, resolvendo conflitos de nome.

        Args:
            source: Caminho do arquivo de origem
            new_name: Novo nome do arquivo (apenas nome, não caminho)

        Returns:
            Caminho final do arquivo renomeado

        Raises:
            FileOperationError: Se o arquivo não existir, sem permissão, ou erro ao renomear
        """
        if not source.exists():
            raise FileOperationError(f"Source file does not exist: {source}")

        if not source.is_file():
            raise FileOperationError(f"Source is not a file: {source}")

        target = source.parent / new_name

        try:
            # Resolver conflito de nome se destino já existe
            if target.exists() and target != source:
                from app.domain.file_naming import generate_unique_filename

                original_target = target
                target = generate_unique_filename(target)
                app_logger.info(
                    "File name conflict resolved",
                    extra={
                        "original_target": str(original_target),
                        "resolved_target": str(target),
                    },
                )

            # Verificar permissão de escrita no diretório
            if not os.access(source.parent, os.W_OK):
                raise FileOperationError(
                    f"No write permission in directory: {source.parent}"
                )

            # Renomear em thread pool
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(None, source.rename, target)

            app_logger.debug(
                "File renamed",
                extra={"source": str(source), "target": str(target)},
            )

            return target

        except FileOperationError:
            raise
        except Exception as e:
            app_logger.error(
                "Failed to rename file",
                extra={
                    "source": str(source),
                    "new_name": new_name,
                    "error": str(e),
                },
            )
            raise FileOperationError(f"Failed to rename {source} to {new_name}: {e}")
