"""
Exportadores de resultados de validação.
"""

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.entities import DocumentFile, DocumentStatus, ValidationResult


class ValidationResultsExporter:
    """Exporta resultados de validação para formatos externos."""

    def export_to_excel(self, result: ValidationResult, output_path: Path) -> None:
        """
        Gera um arquivo Excel com os resultados da validação.
        Duas abas: "Validados" e "Não Reconhecidos".
        """
        workbook = openpyxl.Workbook()

        # Aba de Validados
        sheet_validated = workbook.active
        sheet_validated.title = "Validados"
        self._fill_validated_sheet(sheet_validated, result.validated_files)

        # Aba de Não Reconhecidos
        sheet_unrecognized = workbook.create_sheet("Não Reconhecidos")
        self._fill_unrecognized_sheet(sheet_unrecognized, result.unrecognized_files)

        workbook.save(output_path)
        workbook.close()

    def _fill_validated_sheet(self, sheet: Worksheet, files: List[DocumentFile]):
        """Preenche a aba de arquivos validados."""
        headers = [
            "Status",
            "Arquivo",
            "Caminho Original",
            "Código",
            "Revisão",
            "Título",
        ]
        self._setup_header(sheet, headers)

        for row_idx, file in enumerate(files, start=2):
            manifest_item = file.associated_manifest_item

            row_data = [
                file.status.value,
                file.path.name,
                str(file.path.absolute()),
                manifest_item.document_code if manifest_item else "",
                manifest_item.revision if manifest_item else "",
                manifest_item.title if manifest_item else "",
            ]

            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        self._apply_styles(sheet)

    def _fill_unrecognized_sheet(self, sheet: Worksheet, files: List[DocumentFile]):
        """Preenche a aba de arquivos não reconhecidos."""
        headers = ["Status", "Arquivo", "Caminho Original", "Tamanho (KB)"]
        self._setup_header(sheet, headers)

        for row_idx, file in enumerate(files, start=2):
            size_kb = round(file.size_bytes / 1024, 2)

            row_data = [
                file.status.value,
                file.path.name,
                str(file.path.absolute()),
                size_kb,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        self._apply_styles(sheet)

    def _setup_header(self, sheet: Worksheet, headers: List[str]):
        """Configura o cabeçalho da planilha."""
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    def _apply_styles(self, sheet: Worksheet):
        """Aplica estilos básicos e ajusta largura das colunas."""
        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            length = max(len(str(cell.value) or "") for cell in column_cells)
            length = min(length, 100)  # Cap width at 100
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = length + 2
