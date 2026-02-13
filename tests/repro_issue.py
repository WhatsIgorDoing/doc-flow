import asyncio
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest

from app.core.interfaces import IFileSystemManager
from app.domain.entities import DocumentFile, DocumentGroup, ManifestItem
from app.infrastructure.template_filler import OpenpyxlTemplateFiller


class MockFileSystemManager:
    async def create_directory(self, path: Path) -> None:
        pass

    async def move_file(self, source: Path, destination: Path) -> None:
        pass

    async def copy_file(self, source: Path, destination: Path) -> None:
        # Simple copy for test using blocking IO wrapped in async if needed,
        # but here we just do it synchronously as it is a mock/test
        import shutil

        shutil.copy(source, destination)

    async def rename_file(self, source: Path, new_name: str) -> Path:
        return source.with_name(new_name)


@pytest.mark.asyncio
async def test_fill_and_save_inserts_rows_correctly(tmp_path):
    # Setup
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"

    # Create template
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="HEADER")
    ws.cell(row=2, column=1, value="FIM")
    wb.save(template_path)
    wb.close()

    # Create data
    # 3 groups
    groups = []
    for i in range(3):
        manifest_item = ManifestItem(
            document_code=f"DOC-{i}",
            revision=f"A",
            title=f"Title {i}",
            metadata={"FORMATO": "A4"},
        )
        doc_file = DocumentFile(
            path=Path(f"doc_{i}.pdf"),
            size_bytes=100,
            associated_manifest_item=manifest_item,
        )
        group = DocumentGroup(document_code=f"DOC-{i}", files=[doc_file])
        groups.append(group)

    # Mock File Manager
    file_manager = MockFileSystemManager()

    # Service
    filler = OpenpyxlTemplateFiller(file_manager)

    # Execute
    await filler.fill_and_save(template_path, output_path, groups)

    # Verify
    assert output_path.exists()
    wb_out = openpyxl.load_workbook(output_path)
    ws_out = wb_out.active

    # Expected:
    # Row 1: HEADER
    # Row 2: DOC-0 ...
    # Row 3: DOC-1 ...
    # Row 4: DOC-2 ...
    # Row 5: FIM

    assert ws_out.cell(row=1, column=1).value == "HEADER"
    assert ws_out.cell(row=2, column=1).value == "DOC-0"
    assert ws_out.cell(row=3, column=1).value == "DOC-1"
    assert ws_out.cell(row=4, column=1).value == "DOC-2"
    assert ws_out.cell(row=5, column=1).value == "FIM"

    wb_out.close()
