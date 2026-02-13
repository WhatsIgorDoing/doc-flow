"""
Testes unitários para ValidationResultsExporter.
"""

from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl
import pytest

from app.domain.entities import ValidationResult, DocumentFile, ManifestItem, DocumentStatus
from app.infrastructure.exporters import ValidationResultsExporter


@pytest.fixture
def exporter():
    return ValidationResultsExporter()


@pytest.fixture
def sample_result():
    manifest_item = ManifestItem(
        document_code="DOC-001",
        revision="A",
        title="Test Document",
        metadata={}
    )

    validated_file = DocumentFile(
        path=Path("docs/DOC-001_A.pdf"),
        size_bytes=1024,
        status=DocumentStatus.VALIDATED,
        associated_manifest_item=manifest_item
    )

    unrecognized_file = DocumentFile(
        path=Path("docs/unknown.pdf"),
        size_bytes=2048,
        status=DocumentStatus.UNRECOGNIZED
    )

    return ValidationResult(
        validated_files=[validated_file],
        unrecognized_files=[unrecognized_file],
        validated_count=1,
        unrecognized_count=1
    )


def test_export_to_excel_creates_file(exporter, sample_result):
    with TemporaryDirectory() as tmp_dir:
        output_path = Path(tmp_dir) / "test_export.xlsx"
        exporter.export_to_excel(sample_result, output_path)

        assert output_path.exists()

        # Verify content
        wb = openpyxl.load_workbook(output_path)
        assert "Validados" in wb.sheetnames
        assert "Não Reconhecidos" in wb.sheetnames

        # Verify Validados sheet
        ws_val = wb["Validados"]
        assert ws_val["A1"].value == "Status"
        assert ws_val["A2"].value == "Validado"
        assert ws_val["D2"].value == "DOC-001"
        assert ws_val["E2"].value == "A"

        # Verify Unrecognized sheet
        ws_unrec = wb["Não Reconhecidos"]
        assert ws_unrec["A1"].value == "Status"
        assert ws_unrec["A2"].value == "Não Reconhecido"
        assert ws_unrec["B2"].value == "unknown.pdf"
        assert ws_unrec["D2"].value == 2.0  # 2048 bytes = 2 KB

        wb.close()
